#!/usr/bin/env python3
"""
工事台帳Excel → GASアプリ用CSV 移行ツール
株式会社ネクスト・ワン

使い方:
  python3 migrate_ledgers.py [フォルダパス]

指定フォルダ内の工事台帳Excelを読み取り、
GASアプリのスプレッドシートに取り込めるCSVを出力する。
"""

import openpyxl
import csv
import re
import os
import sys
import uuid
from datetime import datetime

# ============================================================
#  ファイル名から台帳番号を抽出
# ============================================================

def extract_project_id_from_filename(filename):
    """ファイル名の【XXXX】パターンから台帳番号を抽出"""
    m = re.search(r'【(\d+)】', filename)
    return m.group(1) if m else None


def extract_status_from_filename(filename):
    """ファイル名からステータス情報を推定"""
    if '入金完了' in filename:
        return '["完工"]'
    if '完工' in filename:
        return '["完工"]'
    if '請求待ち' in filename:
        return '["保留金請求待ち"]'
    if '着工' in filename:
        return '["進行中"]'
    return '["進行中"]'


# ============================================================
#  台帳Excelパーサー
# ============================================================

class LedgerParser:
    """工事台帳ＮＥＷシートをパースする"""

    def __init__(self, filepath):
        self.filepath = filepath
        self.filename = os.path.basename(filepath)
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.ws = self._find_ledger_sheet()
        self.project_id = extract_project_id_from_filename(self.filename)

    def _find_ledger_sheet(self):
        """工事台帳シートを見つける"""
        for name in self.wb.sheetnames:
            if '工事台帳' in name:
                return self.wb[name]
        # 最初のシートにフォールバック
        return self.wb[self.wb.sheetnames[0]]

    def cell(self, row, col):
        """セル値を取得（None安全）"""
        v = self.ws.cell(row, col).value
        return v

    def str_cell(self, row, col):
        """セル値を文字列で取得"""
        v = self.cell(row, col)
        if v is None:
            return ''
        return str(v).strip()

    def num_cell(self, row, col):
        """セル値を数値で取得"""
        v = self.cell(row, col)
        if v is None:
            return 0
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return 0

    # --------------------------------------------------------
    #  ヘッダー情報 → projects
    # --------------------------------------------------------

    def _find_row_by_label(self, label, col=1, start=1, end=50):
        """指定ラベルを含む行を探す（全角スペース等を無視して比較）"""
        normalized_label = re.sub(r'[\s\u3000]+', '', label)
        for row in range(start, end + 1):
            val = self.str_cell(row, col)
            if val:
                normalized_val = re.sub(r'[\s\u3000]+', '', val)
                if normalized_label in normalized_val:
                    return row
        return None

    def parse_header(self):
        """プロジェクトヘッダー情報を抽出"""
        project_id = self.str_cell(2, 8) or self.project_id  # H2

        # 工期の解析
        period_start = self._parse_period_date(
            self.str_cell(7, 2), self.str_cell(7, 3))  # B7, C7
        period_end = self._parse_period_date(
            self.str_cell(7, 6), self.str_cell(7, 7))  # F7, G7

        # 受注金額: "受注金額" ラベル行のG列を探す（固定行番号に依存しない）
        contract_base = 0
        contract_tax = 0
        contract_total = 0
        juchu_row = self._find_row_by_label('受注金額', col=1, start=8, end=40)
        if juchu_row:
            contract_base = self.num_cell(juchu_row, 7)      # G列: 税抜き
            contract_tax = self.num_cell(juchu_row + 1, 7)    # 次行: 消費税
            contract_total = self.num_cell(juchu_row + 2, 7)  # 次々行: 合計

        # 本体から税を再計算（税額セルが空の場合）
        if contract_base and not contract_tax:
            contract_tax = round(contract_base * 0.1)
        if contract_base and not contract_total:
            contract_total = contract_base + contract_tax

        # 目標金額: "目標金額" ラベル行のG列を探す
        target_base = 0
        target_tax = 0
        target_total = 0
        target_rate = 78  # デフォルト
        mokuhyo_row = self._find_row_by_label('目標金額', col=1, start=8, end=40)
        if mokuhyo_row:
            target_base = self.num_cell(mokuhyo_row, 7)
            target_tax = self.num_cell(mokuhyo_row + 1, 7)
            target_total = self.num_cell(mokuhyo_row + 2, 7)
            # 目標粗利率（ラベルから%を抽出）
            target_rate_text = self.str_cell(mokuhyo_row, 1)
            m = re.search(r'(\d+)[%％]', target_rate_text)
            if m:
                target_rate = int(m.group(1))

        return {
            '台帳番号': str(project_id),
            '客先名': self.str_cell(3, 2),         # B3
            '営業担当': self.str_cell(4, 3),        # C4
            '案件名': self.str_cell(5, 2),          # B5
            '住所': self.str_cell(6, 2),            # B6
            '工期開始': period_start,
            '工期終了': period_end,
            '契約金額_本体': contract_base,
            '契約金額_消費税': contract_tax,
            '契約金額_税込': contract_total,
            '目標粗利率': target_rate,
            '目標金額_本体': target_base,
            '目標金額_消費税': target_tax,
            '目標金額_税込': target_total,
            'ステータス': extract_status_from_filename(self.filename),
            'アラートメッセージ': '',
            '作成日時': datetime.now().isoformat(),
            '更新日時': datetime.now().isoformat(),
        }

    def _parse_period_date(self, reiwa_part, date_part):
        """'R7' + ''3/7' → '2025-03-07' のような和暦→西暦変換"""
        reiwa_part = str(reiwa_part or '').strip()
        date_part = str(date_part or '').strip().lstrip("''/")

        # Reiwa year extraction
        m = re.search(r'[RrＲ]?(\d+)', reiwa_part)
        if not m:
            return ''
        reiwa_year = int(m.group(1))
        western_year = reiwa_year + 2018

        # Month/Day extraction
        m2 = re.search(r"'?(\d+)/(\d+)", date_part)
        if m2:
            month = int(m2.group(1))
            day = int(m2.group(2))
            return f'{western_year}-{month:02d}-{day:02d}'

        m3 = re.search(r"'?(\d+)", date_part)
        if m3:
            month = int(m3.group(1))
            return f'{western_year}-{month:02d}-01'

        return ''

    # --------------------------------------------------------
    #  契約変更履歴 → contract_changes
    # --------------------------------------------------------

    def parse_contract_changes(self):
        """契約変更履歴セクション（行8-15付近）をパース"""
        changes = []
        project_id = self.str_cell(2, 8) or self.project_id

        for row in range(10, 26):
            label = self.str_cell(row, 1)
            if not label:
                continue
            if label == '計':
                break

            amount_base = self.num_cell(row, 3)   # C列: 税抜き
            amount_total = self.num_cell(row, 5)   # E列: 税込
            date_str = self.str_cell(row, 2)       # B列: 日付
            note = self.str_cell(row, 7)           # G列: 備考

            if amount_base == 0 and amount_total == 0:
                continue

            changes.append({
                'ID': str(uuid.uuid4()),
                '台帳番号': str(project_id),
                '変更種別': label,
                '変更日': date_str,
                '変更金額_本体': amount_base,
                '変更金額_税込': amount_total,
                '備考': note,
                '作成日時': datetime.now().isoformat(),
            })

        return changes

    # --------------------------------------------------------
    #  月別経費 → expenses
    # --------------------------------------------------------

    def parse_expenses(self):
        """月別経費セクションをパース"""
        expenses = []
        project_id = self.str_cell(2, 8) or self.project_id

        # "業者名" を含むヘッダー行を探す
        expense_start_row = None
        for row in range(28, 50):
            val = self.str_cell(row, 1)
            if val and ('業' in val and '名' in val):
                expense_start_row = row + 1
                break

        if not expense_start_row:
            return expenses

        # "請求日" を含む行（請求入金セクション）を探す → 経費セクションの終端
        billing_header_row = self.ws.max_row
        for row in range(expense_start_row, self.ws.max_row + 1):
            val = self.str_cell(row, 1)
            if val == '請求日':
                billing_header_row = row
                break

        # 走査
        current_year = None
        current_month = None
        prev_vendor = ''

        for row in range(expense_start_row, billing_header_row):
            a_val = self.str_cell(row, 1)

            # 年度行の検出: "R7年度", "R 7年度", "R8年度"
            if a_val and re.search(r'[RrＲ]\s*(\d+)\s*年度', a_val):
                m = re.search(r'[RrＲ]\s*(\d+)', a_val)
                if m:
                    current_year = int(m.group(1)) + 2018
                continue

            # 月ヘッダーの検出: ＜3月分＞, 《4月分》, ＜１月分＞ など
            row_text = a_val
            if not row_text:
                # B列にもヘッダーがある場合がある
                row_text = self.str_cell(row, 2)

            if row_text:
                month_match = re.search(r'[＜《<]?\s*(\d+|[０-９]+)\s*月分\s*[＞》>]?', str(row_text))
                if month_match:
                    month_str = month_match.group(1)
                    # 全角→半角
                    month_str = month_str.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
                    current_month = int(month_str)

                    # 年度が未設定の場合、最初の月から推定
                    if current_year is None:
                        # 3月以降が最初に出現したら令和7年(2025)と仮定
                        current_year = 2025

                    continue

            # 小計・計の行はスキップ
            if a_val and a_val in ('小計', '計'):
                continue

            # 経費行の判定: D列（金額）に値があるか
            amount = self.num_cell(row, 4)
            if amount == 0:
                continue
            if current_month is None:
                continue

            # 業者名
            vendor = a_val
            if vendor == '〃':
                # 備考から適切なサフィックスを推定
                note_combined = ' '.join(filter(None, [self.str_cell(row, 7), self.str_cell(row, 9)]))
                if '値引' in note_combined:
                    vendor = prev_vendor + '(値引)'
                elif note_combined:
                    # 備考の先頭10文字をサフィックスに
                    short_note = note_combined[:10].strip()
                    vendor = prev_vendor + '(' + short_note + ')'
                else:
                    vendor = prev_vendor + '(2)'
            elif vendor:
                prev_vendor = vendor
            else:
                continue  # 業者名なしはスキップ

            # 年の確定: 年度の切り替わりを考慮
            year = current_year
            if year is None:
                year = 2025

            # 月初日フォーマット
            month_date = f'{year}-{current_month:02d}-01'

            # 相殺
            offset = self.str_cell(row, 6) or ''

            # 備考（G列 + I列）
            note_g = self.str_cell(row, 7)
            note_i = self.str_cell(row, 9)
            note = ' '.join(filter(None, [note_g, note_i]))

            expenses.append({
                'ID': str(uuid.uuid4()),
                '台帳番号': str(project_id),
                '月': month_date,
                '仕入先': vendor,
                '金額': amount,
                '相殺': offset if offset != 'なし' else '',
                '備考': note,
                '作成日時': datetime.now().isoformat(),
            })

        return expenses

    # --------------------------------------------------------
    #  請求入金 → billings
    # --------------------------------------------------------

    def parse_billings(self):
        """請求入金セクションをパース"""
        billings = []
        project_id = self.str_cell(2, 8) or self.project_id

        # "請求日" ヘッダー行を探す
        header_row = None
        for row in range(1, self.ws.max_row + 1):
            if self.str_cell(row, 1) == '請求日':
                header_row = row
                break

        if not header_row:
            return billings

        current_billing_date = ''

        for row in range(header_row + 1, self.ws.max_row + 1):
            a_val = self.str_cell(row, 1)
            b_val = self.str_cell(row, 2)
            c_val = self.num_cell(row, 3)  # 請求金額
            e_val = self.str_cell(row, 5)  # 入金予定日
            f_val = self.str_cell(row, 6)  # 手形・相殺
            g_val = self.cell(row, 7)      # 入金確認額
            j_val = self.str_cell(row, 10) # 確認者

            # 計の行で終了
            if a_val == '計':
                break

            # 契約金額の行はスキップ
            if b_val == '契約金額':
                break

            # 相殺行はスキップ（別途メモする場合は備考に入れるが、billingデータとしては不要）
            if f_val and '相殺' in f_val and not b_val:
                continue

            # 請求日の更新
            if a_val:
                current_billing_date = a_val

            # 種別（本工事、保留金、追加工事等）
            category = b_val
            if not category:
                continue

            # 入金確認額
            confirmed = 0
            if g_val is not None:
                try:
                    confirmed = int(float(g_val))
                except (ValueError, TypeError):
                    confirmed = 0

            if c_val == 0 and confirmed == 0:
                continue

            billings.append({
                'ID': str(uuid.uuid4()),
                '台帳番号': str(project_id),
                '請求日': current_billing_date,
                '種別': category,
                '請求金額': c_val,
                '入金予定日': e_val,
                '入金確認額': confirmed,
                '確認者': j_val,
                '作成日時': datetime.now().isoformat(),
            })

        return billings

    def parse_all(self):
        """全セクションをパースして返す"""
        return {
            'project': self.parse_header(),
            'contract_changes': self.parse_contract_changes(),
            'expenses': self.parse_expenses(),
            'billings': self.parse_billings(),
        }


# ============================================================
#  フォルダ走査
# ============================================================

def find_ledger_files(folder_path):
    """フォルダ内の台帳Excelファイルを列挙"""
    files = []
    for f in os.listdir(folder_path):
        if not f.endswith('.xlsx'):
            continue
        if f.startswith('~$'):  # Excel一時ファイル
            continue
        project_id = extract_project_id_from_filename(f)
        if project_id:
            files.append({
                'path': os.path.join(folder_path, f),
                'filename': f,
                'project_id': project_id,
            })
    return sorted(files, key=lambda x: x['project_id'])


# ============================================================
#  CSV出力
# ============================================================

PROJECT_HEADERS = [
    '台帳番号', '客先名', '営業担当', '案件名', '住所',
    '工期開始', '工期終了',
    '契約金額_本体', '契約金額_消費税', '契約金額_税込',
    '目標粗利率', '目標金額_本体', '目標金額_消費税', '目標金額_税込',
    'ステータス', 'アラートメッセージ', '作成日時', '更新日時'
]

CHANGE_HEADERS = [
    'ID', '台帳番号', '変更種別', '変更日', '変更金額_本体', '変更金額_税込', '備考', '作成日時'
]

EXPENSE_HEADERS = [
    'ID', '台帳番号', '月', '仕入先', '金額', '相殺', '備考', '作成日時'
]

BILLING_HEADERS = [
    'ID', '台帳番号', '請求日', '種別', '請求金額', '入金予定日', '入金確認額', '確認者', '作成日時'
]


def write_csv(filepath, headers, rows):
    """CSVファイルを書き出す"""
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


# ============================================================
#  メイン
# ============================================================

def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else '/home/user/nextone'
    output_dir = os.path.join(folder, 'migration_output')
    os.makedirs(output_dir, exist_ok=True)

    files = find_ledger_files(folder)
    if not files:
        print('台帳Excelファイルが見つかりません。')
        print(f'フォルダ: {folder}')
        print('ファイル名に【数字】を含むExcelファイルを配置してください。')
        return

    print(f'=== 工事台帳 移行ツール ===')
    print(f'フォルダ: {folder}')
    print(f'検出ファイル: {len(files)}件')
    print()

    all_projects = []
    all_changes = []
    all_expenses = []
    all_billings = []
    errors = []

    for fi in files:
        project_id = fi['project_id']
        print(f'  [{project_id}] {fi["filename"][:50]}...')
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                parser = LedgerParser(fi['path'])
                data = parser.parse_all()

            p = data['project']
            exp = data['expenses']
            bil = data['billings']
            chg = data['contract_changes']

            all_projects.append(p)
            all_changes.extend(chg)
            all_expenses.extend(exp)
            all_billings.extend(bil)

            exp_total = sum(e['金額'] for e in exp)
            print(f'    案件名: {p["案件名"]}')
            print(f'    契約金額: {p["契約金額_税込"]:,.0f}円')
            print(f'    経費: {len(exp)}件 (合計: {exp_total:,.0f}円)')
            print(f'    請求入金: {len(bil)}件')
            print(f'    契約変更: {len(chg)}件')
            print()

        except Exception as e:
            errors.append((fi['filename'], str(e)))
            print(f'    ❌ エラー: {e}')
            print()

    # CSV出力
    write_csv(os.path.join(output_dir, 'projects.csv'), PROJECT_HEADERS, all_projects)
    write_csv(os.path.join(output_dir, 'contract_changes.csv'), CHANGE_HEADERS, all_changes)
    write_csv(os.path.join(output_dir, 'expenses.csv'), EXPENSE_HEADERS, all_expenses)
    write_csv(os.path.join(output_dir, 'billings.csv'), BILLING_HEADERS, all_billings)

    # サマリー
    print('=' * 50)
    print(f'完了!')
    print(f'  案件数:     {len(all_projects)}件')
    print(f'  契約変更:   {len(all_changes)}件')
    print(f'  経費明細:   {len(all_expenses)}件')
    print(f'  請求入金:   {len(all_billings)}件')
    if errors:
        print(f'  エラー:     {len(errors)}件')
        for fname, err in errors:
            print(f'    - {fname[:40]}: {err}')
    print(f'\n出力先: {output_dir}/')
    print(f'  projects.csv, contract_changes.csv, expenses.csv, billings.csv')


if __name__ == '__main__':
    main()
